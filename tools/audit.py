"""
audit.py — Static code audit tool for the City Sim engine.

Reads engine source files and extracts all entity attributes, actions,
traits, formulas, and constants into CSVs + a single AUDIT.md reference.

Usage:
    python city_sim_Project/tools/audit.py

Outputs (in city_sim_Project/tools/output/):
    entity_attributes.csv   — Every field on every dataclass
    actions.csv             — All actions with mechanics
    traits.csv              — All traits with mechanical effects
    formulas.csv            — All formula functions
    constants.csv           — All constant tables
    AUDIT.md                — Combined reference for Claude
"""
from __future__ import annotations

import ast
import csv
import os
import re
import sys
import textwrap
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Paths ────────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
ENGINE_DIR = PROJECT_DIR / "backend" / "engine"
OUTPUT_DIR = SCRIPT_DIR / "output"

MODELS_PY = ENGINE_DIR / "models.py"
FORMULAS_PY = ENGINE_DIR / "formulas.py"
ACTIONS_DIR = ENGINE_DIR / "actions"
UNIT_ACTIONS_PY = ACTIONS_DIR / "unit.py"
FACTION_ACTIONS_PY = ACTIONS_DIR / "faction.py"
MEMBERSHIP_ACTIONS_PY = ACTIONS_DIR / "membership.py"
NPC_DIR = ENGINE_DIR / "npc"
WEIGHTS_PY = NPC_DIR / "weights.py"


# ── AST Helpers ──────────────────────────────────────────────────────────────

def _parse_file(path: Path) -> ast.Module:
    return ast.parse(path.read_text(encoding="utf-8"), filename=str(path))


def _annotation_str(node) -> str:
    """Best-effort string from an AST annotation node."""
    if node is None:
        return ""
    if isinstance(node, ast.Constant):
        return str(node.value)
    if isinstance(node, ast.Name):
        return node.id
    if isinstance(node, ast.Attribute):
        return f"{_annotation_str(node.value)}.{node.attr}"
    if isinstance(node, ast.Subscript):
        return f"{_annotation_str(node.value)}[{_annotation_str(node.slice)}]"
    if isinstance(node, ast.Tuple):
        return ", ".join(_annotation_str(e) for e in node.elts)
    if isinstance(node, ast.BinOp):
        return f"{_annotation_str(node.left)} | {_annotation_str(node.right)}"
    return ast.dump(node)


def _default_str(node) -> str:
    """Best-effort string from a default-value AST node."""
    if node is None:
        return ""
    if isinstance(node, ast.Constant):
        return repr(node.value)
    if isinstance(node, ast.Name):
        return node.id
    if isinstance(node, ast.Call):
        func = _annotation_str(node.func)
        if func == "field":
            for kw in node.keywords:
                if kw.arg == "default_factory":
                    return f"field(default_factory={_annotation_str(kw.value)})"
            return "field(...)"
        return f"{func}()"
    if isinstance(node, ast.UnaryOp) and isinstance(node.op, ast.USub):
        return f"-{_default_str(node.operand)}"
    if isinstance(node, ast.List):
        return "[]"
    if isinstance(node, ast.Dict):
        return "{}"
    return ""


# ── 1. Entity Attributes ────────────────────────────────────────────────────

# Hand-annotated metadata for fields that need range/persistence info.
# Key = "ClassName.field_name"
_FIELD_META: Dict[str, Tuple[str, str, str]] = {
    # (range, persistence, description)
    # Unit
    "Unit.id":                  ("", "persistent", "Unique identifier"),
    "Unit.name":                ("", "persistent", "Display name"),
    "Unit.faction_1":           ("", "persistent", "Primary faction slot"),
    "Unit.faction_2":           ("", "persistent", "Secondary faction slot"),
    "Unit.faction_3":           ("", "persistent", "Tertiary faction slot"),
    "Unit.is_leader":           ("", "persistent", "Cached — true if leading any faction"),
    "Unit.traits":              ("1-5 tags", "persistent", "Behavioral/mechanical trait tags"),
    "Unit.edge":                ("0-20", "persistent", "Offensive stat; edge+grit=20"),
    "Unit.grit":                ("0-20", "persistent", "Defensive stat; edge+grit=20"),
    "Unit.health":              ("1-100", "persistent", "Creation: 60-95; decays naturally"),
    "Unit.domains":             ("2-5 entries", "persistent", "Domain expertise ratings"),
    "Unit.focus_1":             ("", "persistent", "Primary psychological focus; decay -0.5/cycle"),
    "Unit.focus_2":             ("", "persistent", "Secondary focus; decay -1.0/cycle"),
    "Unit.is_npc":              ("", "persistent", "True if NPC-controlled"),
    "Unit.spy_gates":           ("", "persistent", "Target IDs successfully spied on"),
    "Unit.unstable_domains":    ("", "persistent", "domain_id -> cycles remaining"),
    "Unit.action_taken":        ("", "cycle-only", "Reset each cycle"),
    "Unit.supported_faction":   ("", "cycle-only", "Faction supported this cycle"),
    "Unit.obscure_levels":      ("", "cycle-only", "domain_id -> obscure defense level"),
    "Unit.cycle_protect_level": ("0-N", "cycle-only", "Protect level taken this cycle"),
    # Faction
    "Faction.id":               ("", "persistent", "Unique identifier"),
    "Faction.name":             ("", "persistent", "Display name"),
    "Faction.domain_primary":   ("", "persistent", "Single primary domain"),
    "Faction.rating":           ("1.0+", "persistent", "Power rating; floor used for capacity"),
    "Faction.traits":           ("", "persistent", "Faction behavioral traits"),
    "Faction.leader_id":        ("", "persistent", "Current leader unit ID"),
    "Faction.leadership_need":  ("0.0-20.0", "persistent", "Accumulates when leaderless"),
    "Faction.entrench":         ("1-100", "persistent", "Organizational resilience"),
    "Faction.member_ids":       ("", "persistent", "Named member unit IDs"),
    "Faction.level_1_count":    ("0+", "persistent", "Background Level 1 members (no tracking)"),
    "Faction.relationships":    ("", "persistent", "Override relationships to other factions"),
    "Faction.floor":            ("", "persistent", "Last confirmed level; auto-init from rating"),
    "Faction.unstable_stacks":  ("0-3", "persistent", "-1 penalty per stack to rolls"),
    "Faction.leaderless_proxy_id": ("", "cycle-only", "Temporary proxy leader this cycle"),
    "Faction.action_taken":     ("", "cycle-only", "Reset each cycle"),
    # Domain
    "Domain.id":                ("", "persistent", "Unique identifier"),
    "Domain.name":              ("", "persistent", "Display name"),
    "Domain.cap":               ("", "persistent", "Total weight supported"),
    "Domain.utilization":       ("0+", "persistent", "Recalculated each cycle"),
    "Domain.drift":             ("", "persistent", "Natural entrench change per cycle"),
    "Domain.relationships":     ("", "persistent", "Domain-to-domain dispositions"),
    # WorldState
    "WorldState.cycle":         ("0+", "persistent", "Current cycle number"),
    "WorldState.chaos":         ("", "persistent", "domain_id -> chaos level 0-10"),
    "WorldState.power_vacuums": ("", "persistent", "{domain_id, cycles_remaining, origin_unit_id}"),
    "WorldState.sm_attention":  ("0+", "persistent", "Social Media attention level"),
    "WorldState.sm_state":      ("", "persistent", "baseline | elevated | crisis"),
    # Sub-objects
    "DomainRating.domain_id":   ("", "persistent", "Which domain"),
    "DomainRating.rating":      ("1.0+", "persistent", "e.g. 4.67; floor() for resolution"),
    "DomainRating.entrench":    ("0.0-1.0", "persistent", "Entrenchment level"),
    "DomainRating.floor":       ("", "persistent", "Last confirmed level; auto-init"),
    "FactionSlot.faction_id":   ("", "persistent", "Which faction"),
    "FactionSlot.inertia":      ("0.0-cap", "persistent", "Bond strength; cap from inertia_cap()"),
    "FocusSlot.target_id":      ("", "persistent", "Focus target"),
    "FocusSlot.target_type":    ("", "persistent", "unit | faction | domain | goal"),
    "FocusSlot.score":          ("0.0-10.0", "persistent", "Focus intensity"),
    "FocusSlot.source_trait":   ("", "persistent", "Trait that generated this focus"),
    "FactionRelationship.faction_id": ("", "persistent", "Target faction"),
    "FactionRelationship.trait": ("", "persistent", "Friend | Foe | Neutral"),
    "DomainRelationship.domain_id": ("", "persistent", "Target domain"),
    "DomainRelationship.trait":  ("", "persistent", "Friend | Foe | Hide | Client | Neutral"),
}


def extract_entity_attributes() -> List[dict]:
    """Parse models.py for all @dataclass fields."""
    tree = _parse_file(MODELS_PY)
    rows: List[dict] = []

    for node in ast.walk(tree):
        if not isinstance(node, ast.ClassDef):
            continue
        # Check for @dataclass decorator
        is_dc = any(
            (isinstance(d, ast.Name) and d.id == "dataclass")
            or (isinstance(d, ast.Attribute) and d.attr == "dataclass")
            for d in node.decorator_list
        )
        if not is_dc:
            continue

        class_name = node.name
        for item in node.body:
            if isinstance(item, ast.AnnAssign) and isinstance(item.target, ast.Name):
                field_name = item.target.id
                type_str = _annotation_str(item.annotation)
                default = _default_str(item.value) if item.value else ""
                key = f"{class_name}.{field_name}"
                meta = _FIELD_META.get(key, ("", "", ""))
                rows.append({
                    "entity": class_name,
                    "field": field_name,
                    "type": type_str,
                    "default": default,
                    "range": meta[0],
                    "persistence": meta[1],
                    "description": meta[2],
                    "line": item.lineno,
                })

    return rows


# ── 2. Actions ───────────────────────────────────────────────────────────────

def _extract_functions(path: Path) -> List[Tuple[str, int, str, List[str]]]:
    """Return [(name, lineno, docstring, arg_names)] for top-level functions."""
    tree = _parse_file(path)
    results = []
    for node in ast.iter_child_nodes(tree):
        if isinstance(node, ast.FunctionDef) and not node.name.startswith("_"):
            doc = ast.get_docstring(node) or ""
            args = [a.arg for a in node.args.args]
            results.append((node.name, node.lineno, doc, args))
    return results


# Action metadata — extracted from reading the code.
# Each entry: (function_name, action_name, category, actor_type, target,
#              requires, roll_type, formula_summary, outcomes, trait_modifiers)
_ACTION_DATA = [
    # Unit actions
    ("resolve_grow_unit", "Grow", "unit", "Unit", "None (self domain)",
     "Domain rating in target domain", "Auto / Contested if domain >=85% util",
     "rating += 1/(2^floor+1)", "success | blocked | contested-fail",
     "Expansionary: +20% if util<80%; Satisfied: -23%"),
    ("resolve_protect", "Protect", "unit", "Unit", "None (self domain)",
     "Domain rating", "No roll",
     "entrench += level*4/100; clears Unstable", "success",
     "None"),
    ("resolve_care", "Care", "unit", "Unit", "None (self)",
     "None", "No roll",
     "health += 4*level (cap 100)", "success",
     "None"),
    ("resolve_passive_spy", "Spy (Passive)", "unit", "Unit", "None (domain scan)",
     "Domain rating", "Roll: effective_rating*0.5 + edge",
     "Detects activity count in domain", "success (info only)",
     "Cautious: triggers at 75% not 50%"),
    ("resolve_targeted_spy", "Spy (Targeted)", "unit", "Unit", "Unit or Faction",
     "Domain rating", "Contest: rating+edge vs target grit+vis_mod",
     "Success adds target to spy_gates", "decisive | partial | fail (20-40% alert)",
     "Anonymous: +grit defense; Paranoid: +1 grit; Cautious: passive at 75%"),
    ("resolve_obscure", "Obscure", "unit", "Unit", "None (self domain)",
     "Domain rating", "No roll",
     "obscure_level = rating*0.5; spy must beat", "success",
     "Paranoid: doubled weight in NPC"),
    ("resolve_block", "Block", "unit", "Unit", "Unit or Faction",
     "Domain rating", "Level comparison (no roll)",
     "block>=target: DECISIVE; block=target-1: PARTIAL; else FAIL",
     "decisive (all blocked) | partial (reduced by 1) | fail",
     "None"),
    ("resolve_harm", "Harm", "unit", "Unit", "Unit or Faction",
     "Spy gate on target; domain rating", "No roll (level comparison)",
     "Stage1: entrench drain 6N/100; Stage2: rating dmg if entrench=0",
     "success | blocked (by Protect)",
     "Protect interaction: protect>=harm blocks; protect<harm reduces"),
    ("resolve_attack", "Attack", "unit", "Unit", "Unit or Faction",
     "Spy gate on target; domain rating", "Contest: rating+edge vs rating+grit+entrench_bonus",
     "Decisive: -1 rating, -8 hp; Partial: Unstable 2 cycles, -3 hp; Fail: attacker exposed",
     "decisive | partial | fail",
     "Fragile: +3 hp damage; Callous: +2 edge"),
    ("resolve_defend", "Defend", "unit", "Unit", "None (self domain)",
     "Domain rating", "Roll: rating + grit + entrench_bonus",
     "Defensive roll logged; used to counter attacks", "success",
     "Cautious: +2 grit"),
    ("resolve_steal", "Steal", "unit", "Unit", "Highest unit in domain",
     "Domain rating", "Contest: rating+edge vs target rating+grit",
     "Decisive: +0.25/-0.25 rating; Partial: +0.10/-0.10; Fail: exposed",
     "decisive | partial | fail",
     "None"),
    # Faction actions
    ("resolve_grow_faction", "Grow (Faction)", "faction", "Faction", "None",
     "None", "No roll",
     "rating += 1/(2^floor+1)", "success",
     "Expansionary: +11%"),
    ("resolve_support_faction", "Support Faction", "faction", "Unit -> Faction", "Faction (must be member)",
     "Faction membership", "No roll",
     "inertia += 0.20; unit set as active supporter", "success | fail (not member)",
     "None"),
    ("resolve_evolve", "Evolve", "faction", "Faction", "None (self)",
     "Has leader", "No roll (25/25/50 probability)",
     "Add leader trait / remove faction trait / both", "success | no_op",
     "None (trait-driven outcome)"),
    # Membership actions
    ("resolve_seek_leadership", "Seek Leadership", "membership", "Unit", "Faction",
     "Faction membership", "Leaderless: roll vs LN; Leader: contest vs grit+entrench",
     "Decisive: immediate transfer; Partial: leader Unstable 2 cycles; Fail: exposed",
     "decisive | partial | fail",
     "Ambitious: +2 edge; Meritocratic: domain rating instead of grit"),
    ("resolve_join", "Join", "membership", "Unit", "Faction",
     "Open faction slot; capacity", "Roll vs DC (base 5)",
     "DC mods: Open -2, Insular +3; Ideological: 50% doctrine check",
     "success | fail",
     "Open: -2 DC; Insular: +3 DC; Ideological: doctrine check"),
    ("resolve_leave", "Leave", "membership", "Unit", "Faction",
     "Faction membership", "No roll",
     "Always succeeds; high inertia (>=6) may trigger retaliation",
     "success",
     "None (inertia-driven consequences)"),
    ("resolve_kick", "Kick", "membership", "Faction", "Unit (lowest member)",
     "Members exist", "No roll (auto-triggered)",
     "Target = lowest inertia/rank/rating member (never leader)",
     "success | fail (no valid target)",
     "Hierarchical: lowest rank; Meritocratic: lowest domain rating"),
    ("resolve_recruit", "Recruit", "membership", "Faction", "None (new L1)",
     "None", "No roll (auto-triggered)",
     "level_1_count += 1", "success",
     "None"),
]


def extract_actions() -> List[dict]:
    """Return action rows from the known action definitions."""
    # Verify functions actually exist in source
    existing_funcs = set()
    for path in [UNIT_ACTIONS_PY, FACTION_ACTIONS_PY, MEMBERSHIP_ACTIONS_PY]:
        if path.exists():
            for name, _, _, _ in _extract_functions(path):
                existing_funcs.add(name)

    rows = []
    for (func, name, cat, actor, target, requires,
         roll_type, formula, outcomes, traits) in _ACTION_DATA:
        source_file = {
            "unit": "actions/unit.py",
            "faction": "actions/faction.py",
            "membership": "actions/membership.py",
        }.get(cat, "?")
        rows.append({
            "action": name,
            "category": cat,
            "function": func,
            "source": source_file,
            "actor_type": actor,
            "target": target,
            "requires": requires,
            "roll_type": roll_type,
            "formula": formula,
            "outcomes": outcomes,
            "trait_modifiers": traits,
            "in_code": "YES" if func in existing_funcs else "MISSING",
        })
    return rows


# ── 3. Traits ────────────────────────────────────────────────────────────────

def _parse_weight_dicts() -> Tuple[Dict[str, Dict[str, float]], Dict[str, Dict[str, float]]]:
    """
    Parse npc/weights.py to extract TRAIT_WEIGHTS and FACTION_TRAIT_WEIGHTS.
    Returns (unit_trait_weights, faction_trait_weights) as {trait: {action: weight}}.
    """
    unit_weights: Dict[str, Dict[str, float]] = {}
    faction_weights: Dict[str, Dict[str, float]] = {}

    if not WEIGHTS_PY.exists():
        return unit_weights, faction_weights

    tree = _parse_file(WEIGHTS_PY)

    for node in ast.iter_child_nodes(tree):
        # Match: TRAIT_WEIGHTS: ... = { ... }  and  FACTION_TRAIT_WEIGHTS: ... = { ... }
        target_name = None
        if isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
            target_name = node.target.id
            value_node = node.value
        elif isinstance(node, ast.Assign):
            for t in node.targets:
                if isinstance(t, ast.Name):
                    target_name = t.id
                    break
            value_node = node.value
        else:
            continue

        if target_name not in ("TRAIT_WEIGHTS", "FACTION_TRAIT_WEIGHTS"):
            continue

        dest = unit_weights if target_name == "TRAIT_WEIGHTS" else faction_weights

        # value_node should be a Dict literal: { "TraitName": { "Action": weight, ... }, ... }
        if not isinstance(value_node, ast.Dict):
            continue

        for key_node, val_node in zip(value_node.keys, value_node.values):
            if not isinstance(key_node, ast.Constant) or not isinstance(val_node, ast.Dict):
                continue
            trait_name = key_node.value
            action_weights: Dict[str, float] = {}
            for ak, av in zip(val_node.keys, val_node.values):
                if isinstance(ak, ast.Constant) and isinstance(av, ast.Constant):
                    action_weights[ak.value] = av.value
            dest[trait_name] = action_weights

    return unit_weights, faction_weights


def _format_weights(weights: Dict[str, float]) -> str:
    """Format {action: weight} dict as a compact string for CSV/markdown."""
    if not weights:
        return ""
    parts = [f"{action} +{int(w)}" for action, w in
             sorted(weights.items(), key=lambda x: -x[1])]
    return "; ".join(parts)


def extract_traits() -> List[dict]:
    """Scan engine source for all trait references and their mechanical effects."""
    # Parse NPC weight tables from source
    unit_weights, faction_weights = _parse_weight_dicts()

    trait_data: Dict[str, dict] = {}

    # Known trait definitions with mechanical effects
    _KNOWN_TRAITS = [
        # Unit traits
        ("Ambitious", "Unit", "+2 edge on Seek Leadership rolls",
         "Seek Leadership"),
        ("Resilient", "Unit", "Natural health decay = 0 (instead of -1)",
         "Health decay"),
        ("Fragile", "Unit", "Natural decay = -2; +3 additional damage from attacks",
         "Health decay, Attack"),
        ("Loner", "Unit", "Inertia caps: -2/-3/-4 per slot; join_desire * 0.2",
         "Inertia, Join desire"),
        ("Joiner", "Unit", "Inertia caps: +2/+3/+4 per slot; join_desire * 1.5",
         "Inertia, Join desire"),
        ("Connected", "Unit", "join_desire * 1.2",
         "Join desire"),
        ("Expansionary", "Unit", "Grow increment +20% if domain utilization < 80%",
         "Grow"),
        ("Satisfied", "Unit", "Grow increment -23% (0.77 multiplier)",
         "Grow"),
        ("Paranoid", "Unit", "+1 grit vs Targeted Spy; Obscure weight doubled in NPC",
         "Spy (Targeted), Obscure"),
        ("Anonymous", "Unit", "+grit (doubled) as defense vs Spy",
         "Spy (Targeted)"),
        ("Cautious", "Unit", "Passive Spy at 75% completion (not 50%); +2 grit on Defend",
         "Spy (Passive), Defend"),
        ("Loyal", "Unit", "+0.10 inertia/cycle in faction_1",
         "Inertia"),
        ("Ideological", "Unit", "Affects Join DC checks (doctrine gate)",
         "Join"),
        ("Hierarchical", "Unit", "Faction governance structure",
         "Kick target selection"),
        ("Callous", "Unit", "+2 edge on Attack rolls",
         "Attack"),
        # Faction traits
        ("Hierarchical", "Faction", "Kick target = lowest rank (last in member_ids)",
         "Kick"),
        ("Meritocratic", "Faction", "Seek Leadership: domain rating instead of grit; Kick: lowest domain rating",
         "Seek Leadership, Kick"),
        ("Open", "Faction", "Join DC -2; Recruit threshold 0.70; Kick threshold 0.90",
         "Join, Recruit, Kick"),
        ("Insular", "Faction", "Join DC +3; Recruit threshold 0.60; Kick threshold 0.75",
         "Join, Recruit, Kick"),
        ("Ideological", "Faction", "Join: 50% doctrine check gate",
         "Join"),
        ("Expansionary", "Faction", "Grow increment +11% (1.11 multiplier)",
         "Grow (Faction)"),
        ("Defensive", "Faction", "Combat bonuses (defensive posture)",
         "Combat"),
        ("Corrupt", "Faction", "Inertia drain -0.05/cycle for all members",
         "Inertia"),
    ]

    for name, applies_to, effect, actions in _KNOWN_TRAITS:
        key = f"{name}_{applies_to}"
        # Look up weight modifiers from the parsed tables
        if applies_to == "Unit":
            weights = unit_weights.get(name, {})
        else:
            weights = faction_weights.get(name, {})

        trait_data[key] = {
            "trait": name,
            "applies_to": applies_to,
            "mechanical_effect": effect,
            "actions_affected": actions,
            "weight_modifiers": _format_weights(weights),
        }

    # Also pick up any traits that are ONLY in the weight tables (not in _KNOWN_TRAITS)
    for wt_name, wt_dict in unit_weights.items():
        key = f"{wt_name}_Unit"
        if key not in trait_data:
            trait_data[key] = {
                "trait": wt_name,
                "applies_to": "Unit",
                "mechanical_effect": "(weights only — no hard-coded mechanic found)",
                "actions_affected": ", ".join(wt_dict.keys()),
                "weight_modifiers": _format_weights(wt_dict),
            }

    for wt_name, wt_dict in faction_weights.items():
        key = f"{wt_name}_Faction"
        if key not in trait_data:
            trait_data[key] = {
                "trait": wt_name,
                "applies_to": "Faction",
                "mechanical_effect": "(weights only — no hard-coded mechanic found)",
                "actions_affected": ", ".join(wt_dict.keys()),
                "weight_modifiers": _format_weights(wt_dict),
            }

    # Verify traits actually appear in source (scan all engine .py files)
    all_source = ""
    for path in ENGINE_DIR.rglob("*.py"):
        all_source += path.read_text(encoding="utf-8")

    rows = []
    for key, data in trait_data.items():
        # Check if trait name appears in source
        in_code = "YES" if f'"{data["trait"]}"' in all_source else "NOT FOUND"
        data["in_code"] = in_code
        rows.append(data)

    return rows


# ── 4. Formulas ──────────────────────────────────────────────────────────────

def extract_formulas() -> List[dict]:
    """Parse formulas.py for all public functions with their docstrings."""
    rows = []
    if not FORMULAS_PY.exists():
        return rows

    tree = _parse_file(FORMULAS_PY)
    source_lines = FORMULAS_PY.read_text(encoding="utf-8").splitlines()

    for node in ast.iter_child_nodes(tree):
        if isinstance(node, ast.FunctionDef) and not node.name.startswith("_"):
            doc = ast.get_docstring(node) or ""
            args = [a.arg for a in node.args.args]
            ret = _annotation_str(node.returns) if node.returns else ""

            # Extract the first line of docstring as summary
            doc_lines = doc.strip().splitlines()
            summary = doc_lines[0] if doc_lines else ""

            # Try to find the core formula expression in the docstring
            formula_expr = ""
            for line in doc_lines:
                line_s = line.strip()
                if any(op in line_s for op in ["=", "×", "*", "/"]) and not line_s.startswith("|"):
                    formula_expr = line_s
                    break

            rows.append({
                "function": node.name,
                "line": node.lineno,
                "args": ", ".join(args),
                "returns": ret,
                "summary": summary,
                "formula": formula_expr,
                "docstring": doc.strip()[:300],
            })

    return rows


# ── 5. Constants ─────────────────────────────────────────────────────────────

def extract_constants() -> List[dict]:
    """Parse formulas.py for module-level constant assignments."""
    rows = []
    if not FORMULAS_PY.exists():
        return rows

    tree = _parse_file(FORMULAS_PY)
    source = FORMULAS_PY.read_text(encoding="utf-8")
    source_lines = source.splitlines()

    def _scan_constants(tree, source_lines, source_label):
        for node in ast.iter_child_nodes(tree):
            # Handle both Assign and AnnAssign (type-annotated constants)
            targets = []
            if isinstance(node, ast.Assign):
                targets = [t for t in node.targets if isinstance(t, ast.Name)]
            elif isinstance(node, ast.AnnAssign) and isinstance(node.target, ast.Name):
                targets = [node.target]

            for target in targets:
                name = target.id
                stripped = name.lstrip("_")
                if stripped and stripped == stripped.upper() and not stripped.startswith("TYPE"):
                    start = node.lineno - 1
                    end = node.end_lineno if node.end_lineno else start + 1
                    value_text = "\n".join(source_lines[start:end]).strip()
                    if len(value_text) > 300:
                        value_text = value_text[:300] + "..."
                    rows.append({
                        "name": name,
                        "line": node.lineno,
                        "source": source_label,
                        "value": value_text,
                    })

    _scan_constants(tree, source_lines, "formulas.py")

    # Also check models.py
    tree2 = _parse_file(MODELS_PY)
    source2 = MODELS_PY.read_text(encoding="utf-8")
    source_lines2 = source2.splitlines()
    _scan_constants(tree2, source_lines2, "models.py")

    return rows


# ── 6. Scan for undocumented trait references ────────────────────────────────

def scan_trait_references() -> List[dict]:
    """Find all string literals that look like trait checks in engine code."""
    pattern = re.compile(r'"([A-Z][a-z]+(?:\s[A-Z][a-z]+)?)"')
    known_non_traits = {
        "Friend", "Foe", "Neutral", "Hide", "Client",
        "Grow", "Protect", "Care", "Harm", "Block", "Attack", "Defend",
        "Steal", "Obscure", "Expose", "Join", "Leave", "Kick", "Recruit",
        "Evolve", "Support Faction", "Seek Leadership",
        "decisive", "partial", "fail", "tie", "success", "blocked", "no_op",
        "baseline", "elevated", "crisis",
        "open", "passive", "contested", "blocked",
        "unit", "faction", "domain", "goal",
        "edge", "grit",
    }

    refs: Dict[str, List[str]] = {}
    for path in ENGINE_DIR.rglob("*.py"):
        text = path.read_text(encoding="utf-8")
        rel = path.relative_to(ENGINE_DIR)
        for m in pattern.finditer(text):
            name = m.group(1)
            if name not in known_non_traits and len(name) > 2:
                refs.setdefault(name, []).append(str(rel))

    rows = []
    for name, files in sorted(refs.items()):
        rows.append({
            "trait_candidate": name,
            "found_in": "; ".join(sorted(set(files))),
        })
    return rows


# ── CSV Writers ──────────────────────────────────────────────────────────────

def _write_csv(path: Path, rows: List[dict], fieldnames: List[str]):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        w.writerows(rows)
    print(f"  wrote {path.name} ({len(rows)} rows)")


# ── Markdown Generator ───────────────────────────────────────────────────────

def _md_table(rows: List[dict], columns: List[str]) -> str:
    """Generate a markdown table from rows."""
    if not rows:
        return "*No data*\n"
    header = "| " + " | ".join(columns) + " |"
    sep = "| " + " | ".join("---" for _ in columns) + " |"
    lines = [header, sep]
    for row in rows:
        cells = []
        for col in columns:
            val = str(row.get(col, "")).replace("|", "\\|").replace("\n", " ")
            if len(val) > 80:
                val = val[:77] + "..."
            cells.append(val)
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines) + "\n"


def generate_audit_md(
    entities: List[dict],
    actions: List[dict],
    traits: List[dict],
    formulas: List[dict],
    constants: List[dict],
    trait_refs: List[dict],
) -> str:
    """Generate the combined AUDIT.md reference."""
    sections = []
    sections.append("# City Sim Engine Audit\n")
    sections.append(f"*Auto-generated by `tools/audit.py`*\n")
    sections.append("---\n")

    # 1. Entity Attributes
    sections.append("## 1. Entity Attributes\n")
    # Group by entity
    entity_groups: Dict[str, List[dict]] = {}
    for row in entities:
        entity_groups.setdefault(row["entity"], []).append(row)

    for ename in ["DomainRating", "FactionSlot", "FocusSlot",
                   "FactionRelationship", "DomainRelationship",
                   "Unit", "Faction", "Domain", "WorldState"]:
        if ename not in entity_groups:
            continue
        sections.append(f"### {ename}\n")
        cols = ["field", "type", "default", "range", "persistence", "description"]
        sections.append(_md_table(entity_groups[ename], cols))

    # 2. Actions
    sections.append("\n## 2. Actions\n")
    for cat in ["unit", "faction", "membership"]:
        cat_rows = [r for r in actions if r["category"] == cat]
        if cat_rows:
            sections.append(f"### {cat.title()} Actions\n")
            cols = ["action", "actor_type", "target", "requires",
                    "roll_type", "formula", "outcomes", "trait_modifiers", "in_code"]
            sections.append(_md_table(cat_rows, cols))

    # 3. Traits
    sections.append("\n## 3. Traits\n")
    sections.append("### Unit Traits\n")
    unit_traits = [r for r in traits if r["applies_to"] == "Unit"]
    cols = ["trait", "mechanical_effect", "actions_affected", "weight_modifiers", "in_code"]
    sections.append(_md_table(unit_traits, cols))

    sections.append("### Faction Traits\n")
    faction_traits = [r for r in traits if r["applies_to"] == "Faction"]
    sections.append(_md_table(faction_traits, cols))

    # 4. Formulas
    sections.append("\n## 4. Formulas\n")
    cols = ["function", "args", "returns", "summary", "formula"]
    sections.append(_md_table(formulas, cols))

    # 5. Constants
    sections.append("\n## 5. Constants\n")
    cols = ["name", "source", "value"]
    sections.append(_md_table(constants, cols))

    # 6. Trait Reference Scan
    sections.append("\n## 6. Trait Reference Scan\n")
    sections.append("*All capitalized string literals found in engine code that may be traits:*\n")
    cols = ["trait_candidate", "found_in"]
    sections.append(_md_table(trait_refs, cols))

    return "\n".join(sections)


# ── XLSX Writer ─────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
_CELL_ALIGN = Alignment(vertical="top", wrap_text=True)
_THIN_BORDER = Border(
    bottom=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
)


def _add_sheet(wb: Workbook, title: str, rows: List[dict], columns: List[str]):
    """Add a formatted sheet to the workbook."""
    ws = wb.create_sheet(title=title)

    # Header row
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = _CELL_ALIGN
            cell.border = _THIN_BORDER

    # Auto-fit column widths (capped at 50)
    for col_idx, col_name in enumerate(columns, 1):
        max_len = len(col_name)
        for row in rows:
            val = str(row.get(col_name, ""))
            max_len = max(max_len, min(len(val), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    # Freeze header row
    ws.freeze_panes = "A2"
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"


def write_xlsx(
    path: Path,
    entities: List[dict],
    actions: List[dict],
    traits: List[dict],
    formulas: List[dict],
    constants: List[dict],
    trait_refs: List[dict],
):
    """Write all audit data to a single .xlsx workbook."""
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)

    _add_sheet(wb, "Entity Attributes", entities,
               ["entity", "field", "type", "default", "range", "persistence", "description", "line"])
    _add_sheet(wb, "Actions", actions,
               ["action", "category", "function", "source", "actor_type", "target",
                "requires", "roll_type", "formula", "outcomes", "trait_modifiers", "in_code"])
    _add_sheet(wb, "Traits", traits,
               ["trait", "applies_to", "mechanical_effect", "actions_affected", "weight_modifiers", "in_code"])
    _add_sheet(wb, "Formulas", formulas,
               ["function", "line", "args", "returns", "summary", "formula", "docstring"])
    _add_sheet(wb, "Constants", constants,
               ["name", "line", "source", "value"])
    _add_sheet(wb, "Trait References", trait_refs,
               ["trait_candidate", "found_in"])

    wb.save(path)
    print(f"  wrote {path.name} ({len(wb.sheetnames)} sheets)")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    print("City Sim Engine Audit")
    print("=" * 40)

    # Validate paths
    missing = []
    for p in [MODELS_PY, FORMULAS_PY, UNIT_ACTIONS_PY,
              FACTION_ACTIONS_PY, MEMBERSHIP_ACTIONS_PY]:
        if not p.exists():
            missing.append(str(p))
    if missing:
        print(f"WARNING: Missing files: {missing}")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Output dir: {OUTPUT_DIR}\n")

    # Extract
    print("Extracting entity attributes...")
    entities = extract_entity_attributes()

    print("Extracting actions...")
    actions = extract_actions()

    print("Extracting traits...")
    traits = extract_traits()

    print("Extracting formulas...")
    formulas = extract_formulas()

    print("Extracting constants...")
    constants = extract_constants()

    print("Scanning trait references...")
    trait_refs = scan_trait_references()

    # Write CSVs
    print("\nWriting CSVs...")
    _write_csv(
        OUTPUT_DIR / "entity_attributes.csv", entities,
        ["entity", "field", "type", "default", "range", "persistence", "description", "line"],
    )
    _write_csv(
        OUTPUT_DIR / "actions.csv", actions,
        ["action", "category", "function", "source", "actor_type", "target",
         "requires", "roll_type", "formula", "outcomes", "trait_modifiers", "in_code"],
    )
    _write_csv(
        OUTPUT_DIR / "traits.csv", traits,
        ["trait", "applies_to", "mechanical_effect", "actions_affected", "weight_modifiers", "in_code"],
    )
    _write_csv(
        OUTPUT_DIR / "formulas.csv", formulas,
        ["function", "line", "args", "returns", "summary", "formula", "docstring"],
    )
    _write_csv(
        OUTPUT_DIR / "constants.csv", constants,
        ["name", "line", "source", "value"],
    )
    _write_csv(
        OUTPUT_DIR / "trait_references.csv", trait_refs,
        ["trait_candidate", "found_in"],
    )

    # Write AUDIT.md
    print("\nGenerating AUDIT.md...")
    md = generate_audit_md(entities, actions, traits, formulas, constants, trait_refs)
    audit_path = OUTPUT_DIR / "AUDIT.md"
    audit_path.write_text(md, encoding="utf-8")
    print(f"  wrote AUDIT.md ({len(md)} chars)")

    # Write combined XLSX
    print("\nGenerating audit.xlsx...")
    write_xlsx(OUTPUT_DIR / "audit.xlsx", entities, actions, traits, formulas, constants, trait_refs)

    # Summary
    print("\n" + "=" * 40)
    print(f"Entities:  {len(entities)} fields across {len(set(r['entity'] for r in entities))} classes")
    print(f"Actions:   {len(actions)} ({sum(1 for a in actions if a['in_code']=='YES')} verified in code)")
    print(f"Traits:    {len(traits)} ({sum(1 for t in traits if t['in_code']=='YES')} verified in code)")
    print(f"Formulas:  {len(formulas)} functions")
    print(f"Constants: {len(constants)} tables")
    print(f"Trait refs: {len(trait_refs)} candidates scanned")
    print("\nDone.")


if __name__ == "__main__":
    main()
